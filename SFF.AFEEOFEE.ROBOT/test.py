import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.drawing.image import Image

__author__ = 'Ika Belerma'

class CustomXlsxLibrary:

    ROBOT_LIBRARY_SCOPE = 'TEST SUITE'

    def __init__(self):
        self.data = []
        self.workbook = None
        self.path = None

    def open_excel_file(self, path_to_excel_file, is_data_only=True):
        """Returns a workbook instance from the an xls file provided.
        Other keywords will ask for an instantiated workbook as an argument.
        Example:
        | Open Excel File | Book1.xls |
        """
        self.path = path_to_excel_file
        self.workbook = load_workbook(path_to_excel_file,data_only=is_data_only)
		
    def create_workbook(self):
        self.workbook = Workbook()

    def get_cell_value_by_sheet_name(self, sheet_name, cell_name):
        """Returns the value of the cell indicated, given the sheet name.
        Example:
        | Open Excel File | Book1.xls |
        | ${A1} | Get Cell Value By Sheet Name | ${workbook} | Sheet1 | A1 |
        This example returns the value of cell "A1" from "Sheet 1" from the xls file "Book1.xls".
        """
        workbook_sheet = self.workbook.get_sheet_by_name(sheet_name)
        cell_value = self._return_cell_value(workbook_sheet, cell_name)
        return cell_value

    def get_cell_value_of_active_sheet(self, cell_name):
        """Returns the value of the cell indicated of the currently active sheet.
        Example:
        | Open Excel File | Book1.xls |
        | ${A1} | Get Cell Value Of Active Sheet | ${workbook} | A1 |
        This example returns the value of cell "A1" from "Sheet 1" from the xls file "Book1.xls".
        """
        workbook_sheet = self.workbook.get_active_sheet()
        cell_value = self._return_cell_value(workbook_sheet, cell_name)
        return cell_value

    def get_sheet_names(self):
        """Returns the names of the sheets of the workbook provided.
        Example:
        | Open Excel File | Book1.xls |
        | ${sheet_names} | Get Sheet Names | ${workbook} |
        Given Book1.xls has three sheets with names Sheet1, Sheet2 and Sheet3:
        
        | Log | ${sheet_names} | # ['Sheet1', 'Sheet2', 'Sheet3'] |
        | Log | ${sheet_names[0]} | # Sheet1 |
        | Log | ${sheet_names[1]} | # Sheet2 |
        | Log | ${sheet_names[2]} | # Sheet3 |
        """
        sheet_names = self.workbook.get_sheet_names()
        return sheet_names

    def get_rows(self, sheet_name):
        """ """
        result=[]
        workbook_sheet = self.workbook.get_sheet_by_name(sheet_name)
        temp = workbook_sheet.iter_rows()
        for row in ws.iter_rows():
            for cell in row:
                result.append(cell.value)
                    yield result
        return result
    
    def get_rows_count(self, sheet_name):
        """
        Example:
        | Open Excel File | Book1.xls |
        | ${count} | Get Rows Count | Sheet1 |
        """
        workbook_sheet = self.workbook.get_sheet_by_name(sheet_name)
        rows_count = workbook_sheet.max_row
        return rows_count

    def get_columns_count(self, sheet_name):
        """
        Example:
        | Open Excel File | Book1.xls |
        | ${count} | Get Columns Count | Sheet1 |
        """
        workbook_sheet = self.workbook.get_sheet_by_name(sheet_name)
        colums_count = workbook_sheet.max_column
        return colums_count

    def put_image_to_cell(self, sheet_name, cell_name, picture_path):
        """
        Example:
        | Open Excel File | Book1.xls |
        | Put Image To Cell | Sheet1 | A1 | C:\\apple.png |
        """
        workbook_sheet = self.workbook.get_sheet_by_name(sheet_name)
        img = Image(picture_path)
        img.anchor(workbook_sheet[cell_name])
        workbook_sheet.add_image(img)

    def put_string_to_cell(self, sheet_name, cell_name, value):
        """
        Example:
        | Open Excel File | Book1.xls |
        | Put String To Cell | Sheet1 | A1 | apple |
        """
        workbook_sheet = self.workbook.get_sheet_by_name(sheet_name)
        workbook_sheet[cell_name] = value

    def add_new_sheet(self, name):
        """
        Example:
        | Open Excel File | Book1.xls |
        | Add New Sheet | Sheet2 |
        """
        workbook_sheet = self.workbook.create_sheet(name)

    def edit_name_of_sheet(self, old_name, new_name):
        """
        Example:
        | Open Excel File | Book1.xls |
        | Edit Name Of Sheet | Sheet1 | Sheet3 |
        """
        workbook_sheet = self.workbook.get_sheet_by_name(old_name)
        workbook_sheet.title = new_name

    def save_excel(self, path):
        """
        Example:
        | Open Excel File | Book1.xls |
        | Save Excel | C:\\Doc.xlsx |
        """
        self.path = path
        self.workbook.save(path)
	
    def save_excel_current_directory(self, filename):
        """
        Example:
        | Open Excel File | Book1.xls |
        | Save Excel Current Directory | Doc1.xlsx |
        """
        workdir = os.getcwd()
        print '*DEBUG* Got fname %s' % filename
        self.workbook.save(os.path.join(workdir, filename))
        return workdir

    @staticmethod
    def _return_cell_value(workbook_sheet, cell_name):
        cell_value = workbook_sheet[cell_name].value
        return cell_value